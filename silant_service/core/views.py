from rest_framework import viewsets
from django.contrib.auth.models import Group
from rest_framework.permissions import IsAuthenticated, AllowAny
from .models import Reference, UserReference, Machine, Maintenance, Claim
from .serializers import MachineSerializer, MachineGuestSerializer, MaintenanceSerializer, ClaimSerializer, ReferenceSerializer, UserReferenceSerializer
from django.shortcuts import render
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter
from rest_framework.viewsets import ReadOnlyModelViewSet, ModelViewSet
from rest_framework.response import Response
from django.http import HttpResponse
import openpyxl
from rest_framework.views import APIView
from rest_framework import status
from rest_framework.permissions import AllowAny
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import logging
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.decorators import api_view
from openpyxl import Workbook

from django.http import JsonResponse
from django.contrib.auth.views import LoginView
from django.contrib.auth import authenticate, login
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt


# Получаем логгер
logger = logging.getLogger(__name__)

logger.setLevel(logging.DEBUG)
handler = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)


class ReferenceViewSet(ReadOnlyModelViewSet):
    """
    ViewSet for viewing and filtering Reference details.
    """
    queryset = Reference.objects.all()
    serializer_class = ReferenceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Optionally filters references by `entity_name` and `name` query parameters.
        """
        queryset = super().get_queryset()

        # Фильтрация по `entity_name`
        entity_name = self.request.query_params.get('entity_name')
        if entity_name:
            queryset = queryset.filter(entity_name=entity_name)

        # Фильтрация по `name`
        name = self.request.query_params.get('name')
        if name:
            queryset = queryset.filter(name=name)

        return queryset


    @action(detail=False, methods=['get'])
    def distinct_entities(self, request):
        """
        Returns a list of distinct `entity_name` values.
        """
        entities = Reference.objects.values_list('entity_name', flat=True).distinct()
        return Response(entities)


class UserReferenceViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing and filtering UserReference details.
    """
    queryset = UserReference.objects.all()
    serializer_class = UserReferenceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Optionally filters UserReferences by `user` query parameter.
        """
        queryset = super().get_queryset()

        # Фильтрация по `user`
        user_id = self.request.query_params.get('user')
        if user_id:
            queryset = queryset.filter(user__id=user_id)

        return queryset

    @action(detail=False, methods=['get'])
    def distinct_users(self, request):
        """
        Returns a list of distinct `user` values (user ids).
        """
        users = UserReference.objects.values_list('user', flat=True).distinct()
        return Response(users)


class MachineViewSet(viewsets.ModelViewSet):
    queryset = Machine.objects.all()
    permission_classes = [AllowAny]

    def get_serializer_class(self):
        if not self.request.user.is_authenticated:
            return MachineGuestSerializer
        return MachineSerializer

    @action(detail=False, methods=['put'], permission_classes=[IsAuthenticated])
    def update_machine(self, request):
        serial_number = request.data.get('serial_number')

        if not serial_number:
            return Response({"detail": "Serial number is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            machine = Machine.objects.get(serial_number=serial_number)
        except Machine.DoesNotExist:
            return Response({"detail": "Machine not found."}, status=status.HTTP_404_NOT_FOUND)

        # Получаем first_name клиента из запроса и находим клиента
        client_data = request.data.get('client')
        if client_data:
            client_first_name = client_data.get('first_name')  # Извлекаем first_name клиента
            if client_first_name:
                try:
                    # Ищем клиента по first_name
                    client = UserReference.objects.get(user__first_name=client_first_name)
                    machine.client = client  # Присваиваем машину найденному клиенту
                except UserReference.DoesNotExist:
                    return Response({"detail": "Client with given first name not found."}, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({"detail": "Client first_name is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Обновляем машину с новым клиентом
        serializer = self.get_serializer(machine, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        serial_number = instance.serial_number

        # Логирование перед удалением
        logger.info(f"Удаление машины с серийным номером {serial_number} и всех связанных записей.")

        # Удаление машины (все связанные записи удаляются автоматически)
        self.perform_destroy(instance)

        return Response(
            {"detail": f"Machine with serial number {serial_number} and related records deleted successfully."},
            status=status.HTTP_204_NO_CONTENT)

    def get_queryset(self):
        user = self.request.user
        serial_number = self.request.query_params.get('serial_number')

        #logger.debug(f"User: {user} (is_authenticated: {user.is_authenticated})")

        queryset = Machine.objects.all()

        # Фильтрация по serial_number
        if serial_number:
            queryset = queryset.filter(serial_number=serial_number)
            #logger.debug(f"Filtered queryset by serial_number: {serial_number}")

        # Логика фильтрации по ролям
        if user.is_authenticated:
            if user.groups.filter(name='manager').exists():
                #logger.debug("User is in 'manager' group, returning all machines.")
                return queryset

            if user.groups.filter(name='client').exists():
                # Получаем объект UserReference для текущего пользователя
                client = UserReference.objects.filter(user=user).first()
                if client:
                    #logger.debug(f"User is in 'client' group, filtering machines by client: {client}")
                    return queryset.filter(client=client)
                else:
                    #logger.error("No associated client found for the user.")
                    raise PermissionDenied("User is not associated with a client.")

            if user.groups.filter(name='service_company').exists():
                service_company = UserReference.objects.filter(user=user).first()
                if service_company:
                    #logger.debug(
                        #f"User is in 'service_company' group, filtering machines by service_company: {service_company}")
                    return queryset.filter(service_company=service_company)
                else:
                    #logger.error("No associated service company found for the user.")
                    raise PermissionDenied("User is not associated with a service company.")

        return queryset


class TableViewSet(ReadOnlyModelViewSet):
    queryset = Machine.objects.all().order_by('serial_number')
    serializer_class = MachineSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]

    def get_queryset(self):
        user = self.request.user
        serial_number = self.request.query_params.get('serial_number')

        # Начальный запрос для машин
        machine_queryset = Machine.objects.all()

        # Фильтрация по serial_number для машин
        if serial_number:
            machine_queryset = machine_queryset.filter(serial_number=serial_number)

        # Логика фильтрации по ролям для машин
        if user.is_authenticated:
            if user.groups.filter(name='manager').exists():
                return machine_queryset

            if user.groups.filter(name='client').exists():
                client = UserReference.objects.filter(user=user).first()
                if client:
                    return machine_queryset.filter(client=client)
                else:
                    raise PermissionDenied("User is not associated with a client.")

            if user.groups.filter(name='service_company').exists():
                service_company = UserReference.objects.filter(user=user).first()
                if service_company:
                    return machine_queryset.filter(service_company=service_company)
                else:
                    raise PermissionDenied("User is not associated with a service company.")
        return machine_queryset

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export_excel(self, request):
        user = request.user

        # Получаем отфильтрованные данные для машин
        machine_queryset = self.get_queryset()

        # Получаем отфильтрованные данные для техобслуживания и заявок
        maintenance_queryset = Maintenance.objects.all()
        claim_queryset = Claim.objects.all()

        # Логика фильтрации по ролям для техобслуживания
        if user.is_authenticated:
            if user.groups.filter(name='manager').exists():
                pass  # Менеджер видит все данные
            elif user.groups.filter(name='client').exists():
                client = UserReference.objects.filter(user=user).first()
                if client:
                    maintenance_queryset = maintenance_queryset.filter(machine__client=client)
                    claim_queryset = claim_queryset.filter(machine__client=client)
                else:
                    raise PermissionDenied("User is not associated with a client.")
            elif user.groups.filter(name='service_company').exists():
                service_company = UserReference.objects.filter(user=user).first()
                if service_company:
                    maintenance_queryset = maintenance_queryset.filter(machine__service_company=service_company)
                    claim_queryset = claim_queryset.filter(machine__service_company=service_company)
                else:
                    raise PermissionDenied("User is not associated with a service company.")

        # Создание Excel-файла
        workbook = Workbook()


        # 1. Лист "Machines"
        sheet1 = workbook.active
        sheet1.title = "Machines"
        self.populate_machine_sheet(sheet1, machine_queryset)

        # 2. Лист "Maintenance"
        sheet2 = workbook.create_sheet(title="Maintenance")
        self.populate_maintenance_sheet(sheet2, maintenance_queryset)

        # 3. Лист "Claims"
        sheet3 = workbook.create_sheet(title="Claims")
        self.populate_claims_sheet(sheet3, claim_queryset)

        # Формирование ответа
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="data_export.xlsx"'
        workbook.save(response)
        return response

    def populate_machine_sheet(self, sheet, queryset):
        """Заполнение листа 'Machines'."""
        headers = [
            "№ п/п", "Модель техники", "Зав. № машины", "Модель двигателя", "Зав. № двигателя",
            "Модель трансмиссии (производитель, артикул)", "Зав. № трансмиссии", "Модель ведущего моста",
            "Зав. № ведущего моста", "Модель управляемого моста", "Зав. № управляемого моста",
            "Дата отгрузки с завода", "Покупатель", "Грузополучатель (конечный потребитель)",
            "Адрес поставки (эксплуатации)", "Комплектация (доп. опции)", "Сервисная компания"
        ]
        self.write_headers(sheet, headers)

        for idx, machine in enumerate(queryset, start=1):
            # Извлекаем данные для клиента и сервисной компании
            client_name = machine.client.user.first_name if machine.client and machine.client.user else '—'
            service_company_name = machine.service_company.user.first_name if machine.service_company and machine.service_company.user else '—'

            sheet.append([
                idx,
                machine.model.name if machine.model else '—',
                machine.serial_number,
                machine.engine_model.name if machine.engine_model else '—',
                machine.engine_serial,
                machine.transmission_model.name if machine.transmission_model else '—',
                machine.transmission_serial,
                machine.drive_axle_model.name if machine.drive_axle_model else '—',
                machine.drive_axle_serial,
                machine.steer_axle_model.name if machine.steer_axle_model else '—',
                machine.steer_axle_serial,
                machine.shipment_date.strftime('%Y-%m-%d') if machine.shipment_date else '',
                client_name,  # Покупатель (Имя клиента)
                machine.consignee if machine.consignee else '—',
                machine.delivery_address if machine.delivery_address else '—',
                machine.configuration if machine.configuration else '—',
                service_company_name  # Сервисная компания (Имя компании)
            ])

    def populate_maintenance_sheet(self, sheet, queryset):
        """Заполнение листа 'Maintenance'."""
        headers = [
            "Зав. № машины", "Вид ТО", "Дата проведения ТО", "Наработка, м/час",
            "№ заказ-наряда", "Дата заказ-наряда", "Организация, проводившая ТО"
        ]
        self.write_headers(sheet, headers)

        for idx, maintenance in enumerate(queryset, start=1):
            serial_number = maintenance.machine.serial_number if maintenance.machine else '—'

            sheet.append([
                serial_number,
                maintenance.maintenance_type.name if maintenance.maintenance_type else '—',
                maintenance.maintenance_date.strftime('%Y-%m-%d') if maintenance.maintenance_date else '—',
                maintenance.operating_hours if maintenance.operating_hours else '—',
                maintenance.order_number if maintenance.order_number else '—',
                maintenance.order_date.strftime('%Y-%m-%d') if maintenance.order_date else '—',
                maintenance.organization.name if maintenance.organization else '—'
            ])

    def populate_claims_sheet(self, sheet, queryset):
        """Заполнение листа 'Claims'."""
        headers = [
            "№ п/п", "Зав. № машины", "Дата отказа", "Наработка, м/час", "Узел отказа",
            "Описание отказа", "Способ восстановления", "Используемые запасные части",
            "Дата восстановления", "Время простоя техники"
        ]
        self.write_headers(sheet, headers)

        for idx, claim in enumerate(queryset, start=1):
            serial_number = claim.machine.serial_number if claim.machine else '—'
            failure_date = claim.failure_date.strftime('%Y-%m-%d') if claim.failure_date else '—'
            recovery_date = claim.recovery_date.strftime('%Y-%m-%d') if claim.recovery_date else '—'
            failed_node = claim.failed_node.name if claim.failed_node else '—'
            failure_description = claim.failure_description if claim.failure_description else '—'
            recovery_method = claim.recovery_method.name if claim.recovery_method else '—'
            spare_parts = claim.spare_parts if claim.spare_parts else '—'
            downtime = claim.downtime if claim.downtime else '—'

            sheet.append([
                idx,
                serial_number,
                failure_date,
                claim.operating_hours if claim.operating_hours else '—',
                failed_node,
                failure_description,
                recovery_method,
                spare_parts,
                recovery_date,
                downtime
            ])

    def write_headers(self, sheet, headers):
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_border = Border(
            left=Side(border_style="thin"), right=Side(border_style="thin"),
            top=Side(border_style="thin"), bottom=Side(border_style="thin")
        )

        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = header_border

        for col in sheet.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            column_letter = col[0].column_letter
            sheet.column_dimensions[column_letter].width = max_length + 2

class MaintenanceViewSet(ModelViewSet):
    queryset = Maintenance.objects.select_related('machine')
    serializer_class = MaintenanceSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()

        # Фильтрация по serial_number, если передан параметр в запросе
        serial_number = self.request.query_params.get('serial_number')
        if serial_number:
            #logger.debug(f"Filtering claims by serial number: {serial_number}")
            queryset = queryset.filter(machine__serial_number=serial_number)

        elif user.groups.filter(name='manager').exists():
            #logger.debug("User is in 'manager' group, returning all maintenance records.")
            return queryset

        elif user.groups.filter(name='client').exists():
            # Получаем объект UserReference для текущего пользователя
            client = UserReference.objects.filter(user=user).first()
            if client:
                #logger.debug(f"User is in 'client' group, filtering maintenance by client: {client}")
                queryset = queryset.filter(machine__client=client)
            else:
                logger.error("No associated client found for the user.")
                raise PermissionDenied("User is not associated with a client.")

        elif user.groups.filter(name='service_company').exists():
            # Получаем объект UserReference для текущего пользователя
            service_company = UserReference.objects.filter(user=user).first()
            if service_company:
                #logger.debug( f"User is in 'service_company' group, filtering maintenance by service company: {service_company}")
                queryset = queryset.filter(machine__service_company=service_company)
            else:
                logger.error("No associated service company found for the user.")
                raise PermissionDenied("User is not associated with a service company.")

        else:
            logger.warning("Unauthenticated user or unsupported group. Returning empty queryset.")
            return queryset.none()



        return queryset

    def update(self, request, *args, **kwargs):
        # Получаем объект Maintenance
        instance = self.get_object()

        # Получаем serial_number из запроса
        serial_number = request.data.get('serial_number')

        if serial_number:
            # Ищем машину по serial_number
            try:
                machine = Machine.objects.get(serial_number=serial_number)
                # Обновляем machine в данных запроса
                request.data['machine'] = machine.id
            except Machine.DoesNotExist:
                return Response({"detail": "Machine with the provided serial_number not found."},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            # Если serial_number не передан, не обновляем поле machine
            request.data['machine'] = instance.machine.id

        # Извлекаем строки для maintenance_type и organization
        maintenance_type_name = request.data.get('maintenance_type')
        organization_name = request.data.get('organization')

        # Получаем или создаем записи в Reference для maintenance_type и organization
        if maintenance_type_name:
            try:
                # Мы теперь передаем имя, а не ID
                maintenance_type = Reference.objects.get(
                    name=maintenance_type_name,
                    entity_name='Вид ТО'
                )
                instance.maintenance_type = maintenance_type  # Привязываем объект, а не ID
            except Reference.DoesNotExist:
                return Response({"detail": f"Maintenance type '{maintenance_type_name}' not found."},
                                status=status.HTTP_404_NOT_FOUND)

        if organization_name:
            try:
                # Мы теперь передаем имя, а не ID
                organization = Reference.objects.get(
                    name=organization_name,
                    entity_name='Организация'
                )
                instance.organization = organization  # Привязываем объект, а не ID
            except Reference.DoesNotExist:
                return Response({"detail": f"Organization '{organization_name}' not found."},
                                status=status.HTTP_404_NOT_FOUND)

        # Создаем и проверяем сериализатор
        serializer = self.get_serializer(instance, data=request.data, partial=True)  # Используем instance
        if serializer.is_valid():
            serializer.save()  # Сохраняем изменения
            return Response(serializer.data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def create(self, request, *args, **kwargs):
        serial_number = request.data.get('serial_number')
        #logger.debug(f"Received serial_number: {serial_number}")
        if not serial_number:
            return Response({"detail": "serial_number is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Ищем машину по serial_number
        try:
            machine = Machine.objects.get(serial_number=serial_number)
        except Machine.DoesNotExist:
            return Response({"detail": "Machine with the provided serial_number not found."},
                             status=status.HTTP_404_NOT_FOUND)

        # Добавляем machine в данные запроса
        request.data['machine'] = machine.id  # Привязываем машину по ID
        #logger.debug(f"Machine ID {machine.id} assigned to maintenance request data")

        # Создаем сериализатор и проверяем данные
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()  # Сохраняем новый объект
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        serial_number = request.query_params.get('serial_number')
        if serial_number:
            Maintenance.objects.filter(machine__serial_number=serial_number).delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        return super().destroy(request, *args, **kwargs)



class ClaimViewSet(viewsets.ModelViewSet):
    queryset = Claim.objects.select_related('machine')
    serializer_class = ClaimSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Фильтрация запросов на основе роли пользователя и serial_number."""
        user = self.request.user
        queryset = super().get_queryset()

        # Фильтрация по serial_number, если передан параметр в запросе
        serial_number = self.request.query_params.get('serial_number')
        if serial_number:
            #logger.debug(f"Filtering claims by serial number: {serial_number}")
            queryset = queryset.filter(machine__serial_number=serial_number)

        # Ограничение по ролям
        if user.groups.filter(name='manager').exists():
            #logger.debug("User is in 'manager' group, returning all claims.")
            return queryset  # Менеджер видит все записи

        elif user.groups.filter(name='client').exists():
            # Получаем объект UserReference для текущего пользователя
            client = UserReference.objects.filter(user=user).first()
            if client:
                #logger.debug(f"User is in 'client' group, filtering claims by client: {client}")
                return queryset.filter(machine__client=client)
            else:
                logger.error("No associated client found for the user.")
                raise PermissionDenied("User is not associated with a client.")

        elif user.groups.filter(name='service_company').exists():
            # Получаем объект UserReference для текущего пользователя
            service_company = UserReference.objects.filter(user=user).first()
            if service_company:
                #logger.debug(f"User is in 'service_company' group, filtering claims by service company: {service_company}")
                return queryset.filter(machine__service_company=service_company)
            else:
                logger.error("No associated service company found for the user.")
                raise PermissionDenied("User is not associated with a service company.")

        # Если пользователь не входит в определенные группы
        logger.warning("Unauthenticated user or unsupported group. Returning empty queryset.")
        return queryset.none()

    def update(self, request, *args, **kwargs):
        # Получаем объект
        instance = self.get_object()

        # Получаем serial_number из запроса
        serial_number = request.data.get('serial_number')

        if serial_number:
            # Ищем машину по serial_number
            try:
                machine = Machine.objects.get(serial_number=serial_number)
                #logger.debug('machine', machine)

                # Обновляем machine в данных запроса
                request.data['machine'] = machine.id
            except Machine.DoesNotExist:
                return Response({"detail": "Machine with the provided serial_number not found."},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            # Если serial_number не передан, не обновляем поле machine
            request.data['machine'] = instance.machine.id

        # Извлекаем строки
        failed_node_name = request.data.get('failed_node')
        recovery_method_name = request.data.get('recovery_method')


        # Получаем или создаем записи в Reference
        if failed_node_name:
            try:
                # Мы теперь передаем имя, а не ID
                failed_node = Reference.objects.get(
                    name=failed_node_name,
                    entity_name='Узел отказа'
                )
                instance.failed_node = failed_node  # Привязываем объект, а не ID
            except Reference.DoesNotExist:
                return Response({"detail": f"Узел отказа '{failed_node_name}' not found."},
                                status=status.HTTP_404_NOT_FOUND)

        if recovery_method_name:
            try:
                # Мы теперь передаем имя, а не ID
                recovery_method = Reference.objects.get(
                    name=recovery_method_name,
                    entity_name='Способ восстановления'
                )
                instance.recovery_method = recovery_method  # Привязываем объект, а не ID
            except Reference.DoesNotExist:
                return Response({"detail": f"Способ восстановления '{recovery_method_name}' not found."},
                                status=status.HTTP_404_NOT_FOUND)

        # Создаем и проверяем сериализатор
        serializer = self.get_serializer(instance, data=request.data, partial=True)  # Используем instance
        if serializer.is_valid():
            serializer.save()  # Сохраняем изменения
            return Response(serializer.data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def create(self, request, *args, **kwargs):
        serial_number = request.data.get('serial_number')
        if not serial_number:
            return Response({"detail": "serial_number is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Ищем машину по serial_number
        try:
            machine = Machine.objects.get(serial_number=serial_number)
        except Machine.DoesNotExist:
            return Response({"detail": "Machine with the provided serial_number not found."},
                             status=status.HTTP_404_NOT_FOUND)

        # Добавляем machine в данные запроса
        request.data['machine'] = machine.id  # Привязываем машину по ID

        # Создаем сериализатор и проверяем данные
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()  # Сохраняем новый объект
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        serial_number = request.query_params.get('serial_number')
        if serial_number:
            Maintenance.objects.filter(machine__serial_number=serial_number).delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        return super().destroy(request, *args, **kwargs)


def main_page(request):
    # Получаем текущего аутентифицированного пользователя
    user = request.user

    # Проверяем, что пользователь аутентифицирован
    if user.is_authenticated:
        user_groups = list(user.groups.values_list('name', flat=True))  # Список групп пользователя
        user_first_name = user.first_name if user.first_name else 'Гость'
    else:
        user_groups = []
        user_first_name = 'Гость'

    logger.info(f'User first name: {user_first_name}, Groups: {user_groups}')

    # Передаем данные в шаблон
    return render(request, 'main.html', {
        'user_first_name': user_first_name,
        'user_groups': user_groups,  # Передача списка групп
    })

def machine_table(request):
    # Получаем все машины из базы данных
    machines = Machine.objects.all()
    return render(request, 'machines_table.html', {'machines': machines})

def user_info_view(request):
    # Получаем текущего аутентифицированного пользователя
    user = request.user

    # Проверяем, что пользователь аутентифицирован
    if user.is_authenticated:
        if user.groups.filter(name='manager').exists():
            # Если менеджер, получаем его имя
            user_first_name = user.first_name  # или можно взять другое поле, например, user.username
        elif user.groups.filter(name='client').exists():
            # Если клиент, получаем имя клиента
            user_first_name = user.first_name  # или какое-то другое поле, например, full name
        elif user.groups.filter(name='service_company').exists():
            # Если сервисная компания, получаем имя компании
            user_first_name = user.first_name  # или используйте поле, которое хранит название компании
        else:
            user_first_name = 'Гость'
    else:
        user_first_name = 'Гость'

    logger.info(f'User first name: {user_first_name}')

    # Передаем данные в шаблон
    return render(request, 'main.html', {'user_first_name': user_first_name})


@method_decorator(csrf_exempt, name='dispatch')  # Отключаем CSRF для AJAX-запросов
class CustomLoginView(LoginView):
    def form_invalid(self, form):
        # Возвращаем JSON при ошибке авторизации
        return JsonResponse({"error": "Неверный логин или пароль."}, status=400)

    def form_valid(self, form):
        # В случае успешной авторизации выполняем стандартное поведение
        user = form.get_user()
        login(self.request, user)
        return JsonResponse({"success": True})  # JSON-ответ при успехе


class MachineUpdateView(APIView):
    def post(self, request):
        user = request.user
        updates = request.data.get('updates', [])

        if not user.is_authenticated:
            return Response({'error': 'Authentication required'}, status=403)

        for update in updates:
            machine_id = update.pop('id', None)
            if not machine_id:
                continue

            try:
                machine = Machine.objects.get(id=machine_id)

                if user.groups.filter(name='manager').exists():
                    pass  # Менеджер может изменять всё
                elif user.groups.filter(name='client').exists():
                    return Response({'error': 'Clients cannot edit machines'}, status=403)
                elif user.groups.filter(name='service_company').exists():
                    return Response({'error': 'Access denied for service_company'}, status=403)

                for key, value in update.items():
                    setattr(machine, key, value)

                machine.save()
            except Machine.DoesNotExist:
                continue

        return Response({'success': True})